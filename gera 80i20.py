import pandas as pd
import xlrd
import os
import openpyxl
import ctypes
import sys  
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import pyperclip

# Declarando a variável global
caminho_salvar = ''

def selecionar_arquivo():
    caminho_arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", ".xls;.xlsx")], initialdir=os.path.expanduser("~"))
    if caminho_arquivo:
        entry_arquivo.delete(0, tk.END)
        entry_arquivo.insert(tk.END, caminho_arquivo)

def selecionar_local():
    global caminho_salvar
    caminho_salvar = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialdir=os.path.expanduser("~"))
    if caminho_salvar:
        entry_local.delete(0, tk.END)
        entry_local.insert(tk.END, caminho_salvar)

def abrir_arquivo():
    if os.path.isfile(caminho_salvar):  # Verificando se o arquivo existe
        os.startfile(caminho_salvar)
    else:
        messagebox.showerror("Erro", "Arquivo não encontrado. Por favor, execute a operação primeiro.")

def carregar_logo():
    # Verifica se estamos executando o código compilado ou o script python
    if getattr(sys, 'frozen', False):
        # Estamos executando o código compilado, muda o caminho para a imagem
        caminho_logo = os.path.join(sys._MEIPASS, 'logo.png')
    else:
        # Estamos executando o script python, o caminho permanece o mesmo
        caminho_logo = os.path.join(os.getcwd(), 'logo.png')

    if os.path.isfile(caminho_logo):
        image = Image.open(caminho_logo)
        image.thumbnail((250, 250), Image.ANTIALIAS)
        logo = ImageTk.PhotoImage(image)
        logo_label.configure(image=logo)
        logo_label.image = logo

def calcular_valores():
    caminho_arquivo = entry_arquivo.get()
    caminho_salvar = entry_local.get()

    if not caminho_arquivo:
        messagebox.showerror("Erro", "Selecione um arquivo Excel.")
        return
    if not caminho_salvar:
        messagebox.showerror("Erro", "Selecione o local para salvar o arquivo.")
        return

    try:
        # Converter o arquivo para o formato XLSX, se necessário
        if caminho_arquivo.endswith('.xls'):
            caminho_arquivo_temp = caminho_arquivo + 'x'  # caminho temporário com extensão .xlsx
            xls_book = xlrd.open_workbook(caminho_arquivo)
            workbook = openpyxl.Workbook()
            for i in range(0, xls_book.nsheets):
                xls_sheet = xls_book.sheet_by_index(i)
                workbook.create_sheet(index=i, title=xls_sheet.name)
                sheet = workbook[xls_sheet.name]
                for row in range(0, xls_sheet.nrows):
                    for col in range(0, xls_sheet.ncols):
                        sheet.cell(row=row+1, column=col+1).value = xls_sheet.cell_value(row, col)
            workbook.save(caminho_arquivo_temp)
            caminho_arquivo = caminho_arquivo_temp
            
        # Ler o arquivo XLSX
        df = pd.read_excel(caminho_arquivo, engine='openpyxl')

        # Verificar se as colunas "QTDE" e "VALOR" estão corretamente preenchidas
        if df[['QTDE', 'VALOR']].isnull().any().any():
            raise ValueError("Algumas células nas colunas 'QTDE' ou 'VALOR' estão vazias.")
        if not df[['QTDE', 'VALOR']].apply(pd.to_numeric, errors='coerce').notnull().all().all():
            raise ValueError("Algumas células nas colunas 'QTDE' ou 'VALOR' não contêm valores numéricos válidos.")

        # Realizar cálculos
        df['20%'] = df['VALOR'] * 0.2
        df['80%'] = df['VALOR'] * 0.8
        df['20% x QTDE'] = df['QTDE'] * df['20%']
        df['80% x QTDE'] = df['QTDE'] * df['80%']

        df.to_excel(caminho_salvar, index=False, float_format='%.0f')

        
        # Função para exibir a tela "Cópia dos Valores e Sucesso"
        def exibir_copia_sucesso():
            janela_copia_sucesso = tk.Toplevel(root)
            janela_copia_sucesso.title("Cópia dos Valores e Sucesso")
            
            # Centralizar a nova janela
            center_window(janela_copia_sucesso)
            
            label_sucesso = ttk.Label(janela_copia_sucesso, text="Os valores foram calculados e as colunas criadas com sucesso!", font=("Segoe UI", 12))
            label_sucesso.pack(pady=10)
            
            # Botão para abrir o arquivo já criado
            button_abrir_arquivo = ttk.Button(janela_copia_sucesso, text="Abrir Arquivo", command=abrir_arquivo)
            button_abrir_arquivo.pack(pady=5)

            label_copia = ttk.Label(janela_copia_sucesso, text="Selecione o valor para copiar:", font=("Segoe UI", 12))
            label_copia.pack(pady=5)
            
            # Função para copiar 20%
            def copiar_20():
                pyperclip.copy('|'.join(map(str, df['20%'])))
                messagebox.showinfo("Cópia", "20% copiado com sucesso!")
                janela_copia_sucesso.focus()  # Focar na janela novamente após a cópia
                
            # Função para copiar 80%
            def copiar_80():
                pyperclip.copy('|'.join(map(str, df['80%'])))
                messagebox.showinfo("Cópia", "80% copiado com sucesso!")
                janela_copia_sucesso.focus()  # Focar na janela novamente após a cópia

            # Botão para copiar 20%
            button_20 = ttk.Button(janela_copia_sucesso, text="Copiar 20%", command=copiar_20)
            button_20.pack(pady=5)
            
            # Botão para copiar 80%
            button_80 = ttk.Button(janela_copia_sucesso, text="Copiar 80%", command=copiar_80)
            button_80.pack(pady=5)
            
            
            # Botão para fechar a janela
            button_fechar = ttk.Button(janela_copia_sucesso, text="Fechar", command=janela_copia_sucesso.destroy)
            button_fechar.pack(pady=10)
        
        # Chamar a função para exibir a tela "Cópia dos Valores e Sucesso"
        exibir_copia_sucesso()

        # Remover o arquivo temporário XLSX, se foi criado
        if caminho_arquivo.endswith('.xlsx') and caminho_arquivo != caminho_salvar:
            os.remove(caminho_arquivo)

    except Exception as e:
        messagebox.showerror("Erro", "Erro ao ler ou escrever o arquivo: {}".format(str(e)))


def abrir_codigo():
    # Aqui você pode adicionar o código para abrir o arquivo do código sem alterar toda a estrutura
    pass


def center_window(window):
    window.update_idletasks()
    width = 850
    height = 320
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

root = tk.Tk()
root.title("Gerar 80 - 20")

style = ttk.Style(root)
style.configure("TButton", font=("Segoe UI", 12), foreground="#000000", background="#ffffff")
style.map("TButton",
          foreground=[('pressed', 'black'), ('active', 'black')],
          background=[('pressed', '!disabled', '#ffffff'), ('active', '#ffffff')])

center_window(root)

frame_logo = ttk.Frame(root)
frame_logo.pack(side=tk.LEFT, padx=10)

logo_label = ttk.Label(frame_logo)
logo_label.pack()
carregar_logo()

frame_arquivo = ttk.Frame(root)
frame_arquivo.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

label_arquivo = ttk.Label(frame_arquivo, text="Selecione o Local do arquivo:", font=("Segoe UI", 12))
label_arquivo.grid(row=0, column=0, padx=5, sticky="w")

entry_arquivo = ttk.Entry(frame_arquivo, font=("Segoe UI", 12))
entry_arquivo.grid(row=0, column=1, padx=5, sticky="ew")

button_arquivo = ttk.Button(frame_arquivo, text="Selecionar", command=selecionar_arquivo)
button_arquivo.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

frame_local = ttk.Frame(root)
frame_local.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

label_local = ttk.Label(frame_local, text="Selecione o local para salvar:", font=("Segoe UI", 12))
label_local.grid(row=0, column=0, padx=5, sticky="w")

entry_local = ttk.Entry(frame_local, font=("Segoe UI", 12))
entry_local.grid(row=0, column=1, padx=5, sticky="ew")

button_local = ttk.Button(frame_local, text="Selecionar", command=selecionar_local)
button_local.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

button_calcular = ttk.Button(root, text="Gerar 80 - 20", command=calcular_valores)
button_calcular.pack(pady=10)

root.resizable(False, False)  # Impede que a janela seja redimensionada

root.mainloop()
